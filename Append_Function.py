import pandas as pd
import openpyxl as ox


def append_df_to_excel(filename, df, sheet_name='Sheet1', start_row=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    # ignore [engine] parameter if it was passed

    if 'engine' in to_excel_kwargs:

        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = ox.load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if start_row is None and sheet_name in writer.book.sheetnames:
            start_row = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        print("File does not exist")
        pass

    if start_row is None:
        start_row = 0

    # write out the new sheet
    df = df.astype('float64', errors='ignore')
    df.to_excel(writer, sheet_name, startrow=start_row, **to_excel_kwargs, header=True, float_format="%.2f")

    # save the workbook
    writer.save()
